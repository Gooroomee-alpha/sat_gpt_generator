import csv

import openpyxl

for i in range(1, 7):
    excel = openpyxl.load_workbook(f"sample_xlsx/SAT_{i}.xlsx")

    for sheet in excel.sheetnames:
        current_sheet = excel[sheet]

        # Create a unique CSV file for each sheet
        csv_file = f"sample/SAT_{i}_{sheet}.csv"

        # Write data from the current sheet to the CSV file
        with open(csv_file, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            for row in current_sheet.iter_rows():
                writer.writerow([cell.value for cell in row])
